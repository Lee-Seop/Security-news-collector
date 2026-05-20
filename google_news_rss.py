import feedparser
import datetime
import pytz
import email.utils
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border,Side
import re

kst = pytz.timezone('Asia/Seoul') # 날짜 한국시간 변환

result=[]
result_main=[] # 주요 언론사
medias=['조선','중앙','동아']
          
#데이터 입력
print("키워드 입력: 1.default(해킹 DDoS 개인정보) 2. 직접입력")
keyword = input()
if keyword == '1':
    keyword = '해킹,DDoS,개인정보'
elif keyword == '2':
    print("키워드를 입력하세요(,(쉼표) 구분자 ex)해킹,DDoS,개인정보)//띄어쓰기 금지! :")
    keyword=input()
keywords=keyword.replace(",","%20OR%20")

print("기간 입력: 1.hour 2. day")
select = input()
if select == '1':
    print('h입력:')
    times = input()
    times = times+'h'
elif select == '2':
    print('d입력:')
    times = input()
    times = times+'d'

#검색 URL
rss_url = f'https://news.google.com/rss/search?q='+keywords+'%20when:'+times+'&hl=ko&gl=KR&ceid=KR%3Ako'

#print(rss_url)

feed = feedparser.parse(rss_url) # rss 데이터 수집

#print(feed)

for entry in feed.entries:
    gmt_time = email.utils.parsedate_to_datetime(entry.published)# 날짜 한국시간 변환
    convert_time = str(gmt_time.astimezone(kst))# 날짜 한국시간 변환
    #print("시간:", convert_time[0:16])
    #print(f"제목: {entry.title}")
    #print(f"링크: {entry.link}")
    #print(f"언론사: {entry.source.title}")
    result.append([convert_time[0:16],entry.title,entry.source.title,entry.link])
    # 주요 언론사 탐색
    for media in medias:
        if re.search(media,entry.source.title):
            result_main.append([convert_time[0:16],entry.title,entry.source.title,entry.link])


#엑셀
df1 = pd.DataFrame(result, columns = ['시간','제목', '언론사', 'link'])
df1 = df1.sort_values(by='시간',ascending=True) #오름차순
workbook = Workbook()
sheet1 = workbook.active
sheet1.title = '전체'
        
for row in sheet1.iter_rows():
    for cell in row:
        cell.font = font
        cell.border =border

#폰트 서식
font = Font(size = 12, color ='000000')
#테두리 서식
side = Side(border_style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)


for row in dataframe_to_rows(df1, index=False, header=True):
    sheet1.append(row)



for row in range(2,len(df1)+2):
        sheet1.row_dimensions[row].height = 18
        link_cell = sheet1.cell(row=row, column=4)  # 'link'는 4번째 열
        link_url = link_cell.value  # URL 가져오기
        link_cell.hyperlink = link_url  # 하이퍼링크 추가
        link_cell.font = Font(underline='single', color='0000FF')  # 하이퍼링크 스타일 지정 (파란색, 밑줄)

sheet1.row_dimensions[1].height = 18
sheet1.column_dimensions['A'].width = 18
sheet1.column_dimensions['B'].width = 75
sheet1.column_dimensions['C'].width = 20
sheet1.column_dimensions['D'].width = 150
        
for row in sheet1.iter_rows():
    for cell in row:
        cell.font = font
        cell.border =border

# 주요 언론사 있을때
if result_main:
    df2 = pd.DataFrame(result_main, columns = ['시간','제목', '언론사', 'link'])
    df2 = df2.sort_values(by='시간',ascending=True)
    sheet2 = workbook.create_sheet('주요 언론사')
    #폰트 서식
    font = Font(size = 12, color ='000000')
    #테두리 서식
    side = Side(border_style='thin', color='000000')
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in dataframe_to_rows(df2, index=False, header=True):
        sheet2.append(row)

    for row in range(2,len(df2)+2):
        sheet2.row_dimensions[row].height = 18
        link_cell = sheet2.cell(row=row, column=4)  # 'link'는 4번째 열
        link_url = link_cell.value  # URL 가져오기
        link_cell.hyperlink = link_url  # 하이퍼링크 추가
        link_cell.font = Font(underline='single', color='0000FF')  # 하이퍼링크 스타일 지정 (파란색, 밑줄)

    sheet2.row_dimensions[1].height = 18
    sheet2.column_dimensions['A'].width = 18
    sheet2.column_dimensions['B'].width = 75
    sheet2.column_dimensions['C'].width = 20
    sheet2.column_dimensions['D'].width = 150

    for row in sheet2.iter_rows():
        for cell in row:
            cell.font = font
            cell.border =border

workbook.save('result.xlsx')

print('끝')
