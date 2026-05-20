from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font,Border,Side,Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time

#무한스크롤 갱신
def scroll(url):
    driver = webdriver.Chrome()#크롬 브라우저 실행
    driver.get(url)#주어지 URL 실행
    cnt = 0

    #현재 페이지 높이, driver.execute_script("Java Script 코드")
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") #스크롤 끝까지 내리기
        
        time.sleep(0.5) # 페이지 로드 시간

        new_height = driver.execute_script("return document.body.scrollHeight")

        # 페이지 높이가 갱신 되지 않으면 break
        if new_height == last_height or cnt == 10:
            break
        
        last_height = new_height

        cnt += 1
    return driver.page_source

title_arr = []
media_arr = []
time_arr = []
url_arr = []
keywords_default = 'ddos | 해킹 | 피싱 | 파밍 | 랜섬웨어 | 금융 | 은행'
keywords = ''
PD = 6
sort = 1

#키워드 기간 입력값 받기

key_input = input('키워드 선택 \n1) 기본(ddos 해킹 피싱 파밍 랜섬웨어 금융 은행)\n2) 직접 입력[공백(spacebar)으로 구분]\n입력하세요:')

if key_input ==  '1':
    keywords = keywords_default
elif key_input == '2':
    keywords_arr = list(input('검색할 키워드를 입력하세요 ex) DDoS 해킹 :').split())
    keywords_input = ' | '.join(keywords_arr)
    keywords = keywords_input
else :
    exit(1)

time_input = input('\n기간 선택\n1) n시간 전\n2) 1일전\n입력하세요:')

if time_input == '1' :
    n = int(input('시간을 입력하세요:'))
    PD += n
elif time_input == '2':
    PD = 4
    sort = 0
else:
    exit(1)


# URL 설정
#url = "https://search.naver.com/search.naver?where=news&query=ddos&sm=tab_opt&sort=1&photo=0&field=0&pd=4&ds=&de=&docid=&related=0&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so%3Add%2Cp%3A1d&is_sug_officeid=0&office_category=0&service_area=0"
url = "https://search.naver.com/search.naver?&where=news&query="+keywords+"&sm=tab_opt&sort="+str(sort)+"&photo=0&field=0&pd="+str(PD)+"&ds=&de=&docid=&related=0&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so%3Add%2Cp%3A1d&is_sug_officeid=0&office_category=0&service_area=0"

# 요청 헤더 설정
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"}

# 웹 페이지 요청
#response = requests.get(url, headers=headers)

html_content = scroll(url)

# 페이지가 제대로 로드되었는지 확인
# HTML 파싱
html = BeautifulSoup(html_content, "html.parser")
    
# 뉴스 추출
title = html.select("div.group_news > ul.list_news > li.bx > div.news_wrap > div.news_area > div.news_contents > a.news_tit")
media = html.select("div.group_news > ul.list_news > li.bx > div.news_wrap > div.news_area > div.news_info > div.info_group > a.info.press")
time  = html.select("div.group_news > ul.list_news > li.bx > div.news_wrap > div.news_area > div.news_info > div.info_group >  span.info")

#텍스트 추출
for news in title:
    url = news.get('href')
    tit = news.get('title')
    url_arr.append(url)
    title_arr.append(tit)
for m in media:
    med = m.get_text(strip=True)
    media_arr.append(med)

for t in time:
    tim = t.get_text(strip=True)
    time_arr.append(tim)
result = [[a , b, c, d] for a,b,c,d in zip(title_arr,media_arr,time_arr,url_arr)]

#엑셀 파일 생성
df = pd.DataFrame(result, columns = ['제목','언론사', '시간', 'link'])

workbook = Workbook()
sheet = workbook.active

#폰트 서식
font = Font(size = 12, color ='000000')

#테두리 서식
side = Side(border_style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)


for row in dataframe_to_rows(df, index=False, header=True):
    sheet.append(row)

for row in sheet.iter_rows():
    for cell in row:
        cell.font = font
        cell.border =border

for row in range(2,len(df)+2):
        sheet.row_dimensions[row].height = 25
        sheet[f'A{row}'].alignment = Alignment(vertical = 'center')
        sheet[f'B{row}'].alignment = Alignment(vertical = 'center')
        sheet[f'C{row}'].alignment = Alignment(vertical = 'center')
        sheet[f'D{row}'].alignment = Alignment(vertical = 'center')
        link_cell = sheet.cell(row=row, column=4)  # 'link'는 4번째 열
        link_url = link_cell.value  # URL 가져오기
        link_cell.hyperlink = link_url  # 하이퍼링크 추가
        link_cell.font = Font(underline='single', color='0000FF')  # 하이퍼링크 스타일 지정 (파란색, 밑줄)

sheet.row_dimensions[1].height = 18
sheet.column_dimensions['A'].width = 68
sheet.column_dimensions['B'].width = 17
sheet.column_dimensions['D'].width = 68
        
workbook.save('data.xlsx')
